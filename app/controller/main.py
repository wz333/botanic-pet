import random
from importlib import import_module
from flask import (
    Blueprint, render_template, jsonify, Response
)
import sys
import os
sys.path.append(os.path.dirname(__file__))

# import camera driver
if os.environ.get('CAMERA'):
    Camera = import_module('camera_' + os.environ['CAMERA']).Camera
else:
    from camera_pi import Camera
    
# import plant monitor
from plant_monitor import PlantMonitor
pm = PlantMonitor()
# import light sensor
import time
import board
import adafruit_tsl2591



# Create sensor object, communicating over the board's default I2C bus
i2c = board.I2C()  # uses board.SCL and board.SDA
# i2c = board.STEMMA_I2C()  # For using the built-in STEMMA QT connector on a microcontroller

# Initialize the sensor.
sensor = adafruit_tsl2591.TSL2591(i2c)

# plant data
import openpyxl

# Load the Excel workbook
excel_file = '/home/pi/botanic-pet/app/controller/final_plant_data.xlsx'
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active





bp = Blueprint('main', __name__)



@bp.route('/')
def index():
    return render_template('main/index.html',
                           title='BotanicPet')
                        

@bp.route('/camera')
def camera():
    return render_template('main/camera.html',
                           title='BotanicPet')
def gen(camera):
    """Video streaming generator function."""
    yield b'--frame\r\n'
    while True:
        frame = camera.get_frame()
        yield b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n--frame\r\n'

@bp.route('/chat')
def chat():
    return render_template('main/chat.html',
                           title='BotanicPet')

@bp.route('/pet_status')
def pet_status():
    row_index=3
    status_code=None
    msg="I'm doing great!"
    status_code=random.choice(['00', '03', '03','03','04'])
    wetness = pm.get_wetness()
    temp_c = pm.get_temp()
    temp_f = round(temp_c*1.8+32,2)
    humidity = int(pm.get_humidity())
    light = int(sensor.lux)

    # light conditions
    if light<int(sheet.cell(row=row_index, column=8).value):
        msg="Zzz"
        status_code="512"
    elif light>int(sheet.cell(row=row_index, column=9).value):
        msg="It's dazzlingly bright!"
        status_code="20"
    
    # temperature conditions
    if temp_f<int(sheet.cell(row=row_index, column=4).value):
        msg="Freezingly COLD!"
        status_code="30"
    elif temp_f>int(sheet.cell(row=row_index, column=5).value):
        msg="Way too hot!"
        status_code="31"
    
    # wetness conditions
    if wetness<int(sheet.cell(row=row_index, column=2).value):
        msg="Water me!"
        status_code="40"
    elif wetness>int(sheet.cell(row=row_index, column=3).value):
        msg="Too much water here!"
        status_code="41"
    
    return jsonify({'data':status_code, 'wet': str(wetness), 'temp_c': str(temp_c), 'temp_f': str(temp_f), 'humi': str(humidity), 'lux': str(light), 'msg': msg})
    # return status_code

@bp.route('/video_feed')
def video_feed():
    """Video streaming route. Put this in the src attribute of an img tag."""
    return Response(gen(Camera()),
                    mimetype='multipart/x-mixed-replace; boundary=frame')

# Close the workbook
# workbook.close()

